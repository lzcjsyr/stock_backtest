#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
主板低TTM PE轮动策略

策略描述：
每月从沪深主板市值≥100亿的股票中，按TTM PE升序排列，
选择前10只股票等权重持仓，次月第一个交易日调仓。

TTM PE = 股价 / TTM EPS
TTM EPS = 最近12个月滚动每股收益
"""

import pandas as pd
import sqlite3
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font
import warnings
import json
from playwright.sync_api import sync_playwright
warnings.filterwarnings('ignore')


class LowTTMPEStrategy:
    def __init__(self, db_path=None):
        """初始化策略"""
        if db_path is None:
            # 自动找到项目根目录下的数据库文件
            script_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(script_dir)
            db_path = os.path.join(project_root, "data", "a_stock", "a_stock_data.db")
        
        self.db_path = db_path
        if not os.path.exists(db_path):
            raise FileNotFoundError(f"数据库文件不存在: {db_path}")
        
        self.conn = sqlite3.connect(db_path)
        
        # 策略参数
        self.strategy_name = "主板低TTM PE轮动策略"
        self.min_market_cap = 100  # 最低市值(亿元)
        self.stock_count = 10      # 持仓数量
        self.transaction_cost = 0.0001  # 手续费率(万分之一)
        
        # 回测参数
        self.start_date = "2020-01-01"
        self.end_date = "2025-06-30"
        self.initial_capital = 1000000  # 初始资金100万
        
        # 结果存储
        self.positions = []  # 持仓记录
        self.nav_history = []  # 净值历史
        
        # 交易日缓存
        self._trading_dates_cache = None
        
    def get_trading_dates(self):
        """获取回测期间的交易日期（带缓存）"""
        if self._trading_dates_cache is None:
            query = """
            SELECT DISTINCT trade_date 
            FROM stock_daily_kline 
            WHERE trade_date >= ? AND trade_date <= ?
            ORDER BY trade_date
            """
            dates_df = pd.read_sql(query, self.conn, params=[self.start_date, self.end_date])
            if len(dates_df) == 0:
                raise ValueError(f"在 {self.start_date} 到 {self.end_date} 期间未找到交易日数据")
            self._trading_dates_cache = dates_df['trade_date'].tolist()
        return self._trading_dates_cache
    
    def get_month_end_dates(self, trading_dates):
        """获取每月最后一个交易日"""
        dates = pd.to_datetime(trading_dates)
        month_ends = []
        
        current_month = None
        for date in dates:
            if current_month != (date.year, date.month):
                if current_month is not None:
                    month_ends.append(prev_date.strftime('%Y-%m-%d'))
                current_month = (date.year, date.month)
            prev_date = date
        
        # 添加最后一个月的月末
        if dates.size > 0:
            month_ends.append(dates[-1].strftime('%Y-%m-%d'))
        
        # 不再丢弃首月：允许从起始月份开始选股
        # 例如回测从 2020-01 开始，则 1 月月末选股，2 月首个交易日建仓
        return month_ends
    
    
    def get_ttm_required_dates(self, base_date):
        """
        获取TTM计算所需的财务报告期日期
        
        Returns:
        --------
        tuple: (required_dates, mode)
            required_dates: 需要查询的报告期日期列表
            mode: 计算模式标识
        """
        bd = str(base_date).replace('-', '')
        base_year = int(bd[:4])
        base_month = int(bd[4:6])

        # 四个安全披露窗口：
        # 10-12月：TTM = (Y Q3) + (Y-1 Annual - Y-1 Q3)
        # 7-9月：TTM = (Y Mid) + (Y-1 Annual - Y-1 Mid)  
        # 4-6月：TTM = (Y Q1) + (Y-1 Annual - Y-1 Q1)
        # 1-3月：TTM = (Y-1 Q3) + (Y-2 Annual - Y-2 Q3)
        if base_month >= 10:
            required_dates = [f"{base_year}0930", f"{base_year-1}1231", f"{base_year-1}0930"]
            mode = 'oct_dec'
        elif base_month >= 7:
            required_dates = [f"{base_year}0630", f"{base_year-1}1231", f"{base_year-1}0630"]
            mode = 'jul_sep'
        elif base_month >= 4:
            required_dates = [f"{base_year}0331", f"{base_year-1}1231", f"{base_year-1}0331"]
            mode = 'apr_jun'
        else:  # 1-3月
            required_dates = [f"{base_year-1}0930", f"{base_year-2}1231", f"{base_year-2}0930"]
            mode = 'jan_mar'
        
        return required_dates, mode
    
    def calculate_ttm_eps_from_data(self, eps_data, required_dates, mode):
        """
        基于已获取的EPS数据计算TTM EPS
        
        Parameters:
        -----------
        eps_data : dict
            财务数据字典，格式：{report_date: eps_value}
        required_dates : list
            需要的报告期日期
        mode : str
            计算模式
            
        Returns:
        --------
        float or None: TTM EPS值，无法计算时返回None
        """
        if len(required_dates) != 3:
            return None
            
        # 检查数据完整性
        for date in required_dates:
            if date not in eps_data or eps_data[date] is None:
                return None
        
        # 根据模式计算TTM EPS
        if mode == 'oct_dec':
            # TTM = Y Q3 + (Y-1 Annual - Y-1 Q3)
            current_q3, prev_annual, prev_q3 = required_dates
            ttm_eps = eps_data[current_q3] + (eps_data[prev_annual] - eps_data[prev_q3])
        elif mode == 'jul_sep':
            # TTM = Y Mid + (Y-1 Annual - Y-1 Mid)
            current_mid, prev_annual, prev_mid = required_dates
            ttm_eps = eps_data[current_mid] + (eps_data[prev_annual] - eps_data[prev_mid])
        elif mode == 'apr_jun':
            # TTM = Y Q1 + (Y-1 Annual - Y-1 Q1)
            current_q1, prev_annual, prev_q1 = required_dates
            ttm_eps = eps_data[current_q1] + (eps_data[prev_annual] - eps_data[prev_q1])
        else:  # jan_mar
            # TTM = Y-1 Q3 + (Y-2 Annual - Y-2 Q3)
            prev_q3, prev2_annual, prev2_q3 = required_dates
            ttm_eps = eps_data[prev_q3] + (eps_data[prev2_annual] - eps_data[prev2_q3])
            
        return ttm_eps if (ttm_eps is not None and ttm_eps > 0) else None

    def calculate_ttm_eps(self, stock_code, base_date):
        """计算单个股票的TTM EPS（采用安全披露窗口）"""
        required_dates, mode = self.get_ttm_required_dates(base_date)
        
        # 查询财务数据
        query = """
        SELECT report_date, value 
        FROM stock_financial_abstract
        WHERE stock_code = ? AND indicator = '基本每股收益'
          AND report_date IN ({})
        """.format(','.join(['?'] * len(required_dates)))
        
        eps_df = pd.read_sql(query, self.conn, params=[stock_code] + required_dates)
        
        if len(eps_df) < len(required_dates):
            return None
            
        eps_data = dict(zip(eps_df['report_date'], eps_df['value']))
        return self.calculate_ttm_eps_from_data(eps_data, required_dates, mode)
    
    def select_stocks(self, selection_date):
        """选股函数：选择TTM PE最低的10只股票"""
        print(f"🔍 {selection_date} 开始选股...")
        
        # 获取候选股票池：主板+市值>100亿
        query = """
        SELECT bi.stock_code, bi.stock_name, 
               bi.total_market_value/100000000 as market_cap_yi,
               kl.close_price
        FROM stock_basic_info bi
        INNER JOIN stock_daily_kline kl ON bi.stock_code = kl.stock_code 
        WHERE (bi.stock_code LIKE '6%' OR bi.stock_code LIKE '0%')
          AND LENGTH(bi.stock_code) = 6
          AND bi.total_market_value >= ?
          AND kl.trade_date = ?
        """
        
        candidates = pd.read_sql(query, self.conn, 
                               params=[self.min_market_cap * 100000000, selection_date])
        
        if len(candidates) == 0:
            print(f"❌ {selection_date} 无可选股票")
            return pd.DataFrame()
        
        # ========= 批量计算 TTM EPS（使用统一逻辑） =========
        required_dates, mode = self.get_ttm_required_dates(selection_date)
        
        codes = candidates['stock_code'].dropna().astype(str).unique().tolist()
        if len(codes) == 0:
            print(f"❌ {selection_date} 无有效股票代码")
            return pd.DataFrame()

        # 批量查询财务数据
        code_placeholders = ','.join(['?'] * len(codes))
        date_placeholders = ','.join(['?'] * len(required_dates))
        eps_sql = f"""
        SELECT stock_code, report_date, value 
        FROM stock_financial_abstract
        WHERE indicator = '基本每股收益'
          AND stock_code IN ({code_placeholders})
          AND report_date IN ({date_placeholders})
        """
        eps_params = list(codes) + list(required_dates)
        eps_df = pd.read_sql(eps_sql, self.conn, params=eps_params)

        if eps_df.empty:
            return self._return_empty_selection(selection_date)

        # 转为透视表格式
        pivot = eps_df.pivot_table(index='stock_code', columns='report_date', values='value', aggfunc='last')
        
        # 检查必需列是否存在并筛选有完整数据的股票
        if not set(required_dates).issubset(set(pivot.columns)):
            return self._return_empty_selection(selection_date)
            
        complete_data = pivot[required_dates].dropna()
        if complete_data.empty:
            return self._return_empty_selection(selection_date)

        # 批量计算TTM EPS
        ttm_eps_list = []
        for stock_code in complete_data.index:
            eps_data = complete_data.loc[stock_code].to_dict()
            ttm_eps = self.calculate_ttm_eps_from_data(eps_data, required_dates, mode)
            if ttm_eps is not None:
                ttm_eps_list.append({'stock_code': stock_code, 'ttm_eps': ttm_eps})

        if not ttm_eps_list:
            return self._return_empty_selection(selection_date)
            
        ttm_eps_df = pd.DataFrame(ttm_eps_list)

        # 合并候选池，计算 TTM PE
        merged = candidates.merge(ttm_eps_df, on='stock_code', how='inner')
        if merged.empty:
            return self._return_empty_selection(selection_date)

        merged['ttm_pe'] = merged['close_price'] / merged['ttm_eps']

        # 与原逻辑一致：展示数值四舍五入后参与排序
        merged['market_cap_yi'] = merged['market_cap_yi'].round(2)
        merged['ttm_eps'] = merged['ttm_eps'].round(3)
        merged['ttm_pe'] = merged['ttm_pe'].round(2)

        selected = merged.nsmallest(self.stock_count, 'ttm_pe')[
            ['stock_code', 'stock_name', 'market_cap_yi', 'close_price', 'ttm_eps', 'ttm_pe']
        ]
        
        print(f"✅ {selection_date} 成功选出 {len(selected)} 只股票")
        print(f"   TTM PE范围: {selected['ttm_pe'].min():.2f} - {selected['ttm_pe'].max():.2f}")
        
        return selected
    
    def run_backtest(self, 
                     min_market_cap: float = 100.0,    # 最低市值(亿)，范围：50-500，默认100
                     stock_count: int = 10,            # 选股数量，范围：5-30，默认10
                     transaction_cost: float = 0.0001, # 手续费率，范围：0-0.005，默认万1
                     start_date: str = "2020-01-01",  # 回测开始日期，格式：YYYY-MM-DD
                     end_date: str = "2025-06-30",    # 回测结束日期，格式：YYYY-MM-DD
                     # 向后兼容的年份参数
                     start_year: int = None,          # 已废弃，请使用start_date
                     end_year: int = None             # 已废弃，请使用end_date
                     ):
        """
        策略回测主函数
        
        Parameters:
        -----------
        min_market_cap : float
            最低市值过滤(亿元)，低于此市值的股票被排除
        stock_count : int
            选股数量，决定持仓股票的数量
        transaction_cost : float
            交易手续费率，每次买卖的成本比例
        start_date : str
            回测开始日期，格式：YYYY-MM-DD
        end_date : str
            回测结束日期，格式：YYYY-MM-DD
        """
        
        # 向后兼容处理（优先使用日期参数）
        if start_year is not None and start_date == "2020-01-01":
            start_date = f"{start_year}-01-01"
        if end_year is not None and end_date == "2025-06-30":
            end_date = f"{end_year}-06-30"
        
        # 更新参数
        self.min_market_cap = min_market_cap
        self.stock_count = stock_count
        self.transaction_cost = transaction_cost
        self.start_date = start_date
        self.end_date = end_date
        
        print(f"🚀 开始回测: {self.strategy_name}")
        print(f"📊 参数设置: 市值>{min_market_cap}亿, 选股{stock_count}只, 手续费{transaction_cost*10000:.1f}‱")
        print(f"⏰ 回测期间: {self.start_date} 至 {self.end_date}")
        
        # 获取交易日期
        trading_dates = self.get_trading_dates()
        month_end_dates = self.get_month_end_dates(trading_dates)
        
        # 初始化
        current_nav = 1.0
        current_positions = pd.DataFrame()
        
        # 记录起始净值
        self.nav_history.append({
            'date': self.start_date,
            'nav': current_nav
        })
        
        
        prev_selection_date = None
        
        for i, selection_date in enumerate(month_end_dates):
            print(f"\n📅 第{i+1}期 选股日期: {selection_date}")
            
            # 选股
            selected_stocks = self.select_stocks(selection_date)
            if selected_stocks.empty:
                # 无法选股时，记录当前净值（保持不变）
                self.nav_history.append({
                    'date': selection_date,
                    'nav': current_nav
                })
                continue
            
            # 计算持仓收益(如果有前期持仓)
            if not current_positions.empty and prev_selection_date:
                period_return = self.calculate_period_return(
                    current_positions, prev_selection_date, selection_date)
                current_nav *= (1 + period_return - self.transaction_cost)
                print(f"📈 期间收益率: {period_return*100:.2f}%, 净值: {current_nav:.4f}")
            
            # 更新持仓
            current_positions = selected_stocks.copy()
            current_positions['weight'] = 1.0 / len(selected_stocks)
            
            print(f"🔄 调仓日期: {selection_date}")
            
            # 记录数据
            self.positions.append({
                'period': i + 1,
                'selection_date': selection_date,
                'rebalance_date': selection_date,  # 简化：直接使用选股日作为调仓日
                'stocks': current_positions,
                'nav': current_nav
            })
            
            # 记录净值
            self.nav_history.append({
                'date': selection_date,
                'nav': current_nav
            })
            
            # 更新上一期选股日期
            prev_selection_date = selection_date
        
        # 计算最后一期收益（从最后一次选股到回测结束日）
        if not current_positions.empty and prev_selection_date and prev_selection_date != self.end_date:
            final_return = self.calculate_period_return(
                current_positions, prev_selection_date, self.end_date)
            current_nav *= (1 + final_return - self.transaction_cost)
            print(f"📈 最后一期收益率: {final_return*100:.2f}%, 净值: {current_nav:.4f}")
            
            # 记录最后净值
            self.nav_history.append({
                'date': self.end_date,
                'nav': current_nav
            })
        
        
        print(f"\n🎯 回测完成!")
        print(f"📊 总收益率: {(current_nav - 1) * 100:.2f}%")
        print(f"📈 最终净值: {current_nav:.4f}")
        
        # 生成报告
        self.generate_reports()
        
    def calculate_period_return(self, positions, start_date, end_date):
        """计算持仓期间收益率"""
        total_return = 0.0
        if positions.empty:
            return total_return

        codes = positions['stock_code'].dropna().astype(str).unique().tolist()
        if len(codes) == 0:
            return total_return

        code_placeholders = ','.join(['?'] * len(codes))
        price_sql = f"""
        SELECT stock_code, trade_date, close_price 
        FROM stock_daily_kline 
        WHERE stock_code IN ({code_placeholders}) AND trade_date IN (?, ?)
        """
        price_params = list(codes) + [start_date, end_date]
        prices = pd.read_sql(price_sql, self.conn, params=price_params)

        if prices.empty:
            return total_return

        pivot = prices.pivot_table(index='stock_code', columns='trade_date', values='close_price', aggfunc='last')
        if start_date not in pivot.columns or end_date not in pivot.columns:
            return total_return

        pivot = pivot[[start_date, end_date]].dropna()
        if pivot.empty:
            return total_return

        pivot['stock_return'] = (pivot[end_date] - pivot[start_date]) / pivot[start_date]
        returns_df = pivot[['stock_return']].reset_index()

        merged = positions[['stock_code', 'weight']].merge(returns_df, on='stock_code', how='inner')
        if merged.empty:
            return total_return

        total_return = float((merged['stock_return'] * merged['weight']).sum())
        return total_return
    
    def generate_reports(self):
        """生成回测报告"""
        timestamp = datetime.now().strftime("%m%d_%H%M")
        
        # 确保results目录在项目根目录下
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(script_dir)
        result_dir = os.path.join(project_root, "results", f"主板低TTM_PE策略_{timestamp}")
        os.makedirs(result_dir, exist_ok=True)
        
        # 1. 生成Excel报告
        self.generate_excel_report(result_dir)
        
        # 2. 生成图表
        self.generate_chart(result_dir)
        
        # 3. 生成README
        self.generate_readme(result_dir)
        
        print(f"📁 报告已生成至: {result_dir}")
    
    def generate_excel_report(self, result_dir):
        """生成Excel报告"""
        wb = Workbook()
        
        # 策略概览工作表
        ws_overview = wb.active
        ws_overview.title = "Strategy_Overview"
        
        # 写入策略基本信息
        basic_info = [
            ["策略名称", self.strategy_name],
            ["策略类型", "价值投资策略"],
            ["回测期间", f"{self.start_date} 至 {self.end_date}"],
            ["调仓频率", "月度调仓"],
            ["初始资金", f"{self.initial_capital:,}元"]
        ]
        
        current_row = 1
        ws_overview[f'A{current_row}'] = "=== 策略基本信息 ==="
        ws_overview[f'A{current_row}'].font = Font(bold=True, size=14)
        current_row += 2
        
        for key, value in basic_info:
            ws_overview[f'A{current_row}'] = key
            ws_overview[f'B{current_row}'] = value
            current_row += 1
        
        # 策略详细描述
        current_row += 2
        ws_overview[f'A{current_row}'] = "=== 策略描述 ==="
        ws_overview[f'A{current_row}'].font = Font(bold=True, size=14)
        current_row += 2
        
        strategy_description = [
            ["选股范围", f"沪深主板股票（6开头+0开头），市值≥{self.min_market_cap}亿元"],
            ["选股指标", "TTM PE（滚动12个月市盈率）"],
            ["选股逻辑", f"按TTM PE升序排列，选择前{self.stock_count}只股票"],
            ["权重分配", "等权重配置，每只股票10%"],
            ["调仓规则", "每月最后一个交易日选股，下月第一个交易日调仓"],
            ["TTM计算", "最近12个月滚动每股收益，确保数据时效性"]
        ]
        
        for key, value in strategy_description:
            ws_overview[f'A{current_row}'] = key
            ws_overview[f'B{current_row}'] = value
            current_row += 1
            
        # 关键参数设置
        current_row += 2
        ws_overview[f'A{current_row}'] = "=== 关键参数 ==="
        ws_overview[f'A{current_row}'].font = Font(bold=True, size=14)
        current_row += 2
        
        parameters = [
            ["最低市值要求", f"{self.min_market_cap}亿元"],
            ["选股数量", f"{self.stock_count}只"],
            ["手续费率", f"{self.transaction_cost*10000:.1f}‱"],
            ["风险控制", "无最大回撤限制"],
            ["股票池范围", "沪深主板（排除ST、停牌股票）"]
        ]
        
        for key, value in parameters:
            ws_overview[f'A{current_row}'] = key
            ws_overview[f'B{current_row}'] = value
            current_row += 1
            
        # 核心业绩指标
        current_row += 2
        ws_overview[f'A{current_row}'] = "=== 业绩指标 ==="
        ws_overview[f'A{current_row}'].font = Font(bold=True, size=14)
        current_row += 2
        
        if self.nav_history:
            total_periods = len(self.positions)
            total_return = (self.nav_history[-1]['nav'] - 1) * 100
            final_nav = self.nav_history[-1]['nav']
            years = (pd.to_datetime(self.end_date) - pd.to_datetime(self.start_date)).days / 365.25
            annual_return = (final_nav ** (1/years) - 1) * 100 if years > 0 else 0
            
            performance = [
                ["总收益率", f"{total_return:.2f}%"],
                ["最终净值", f"{final_nav:.4f}"],
                ["年化收益率", f"{annual_return:.2f}%"],
                ["调仓次数", f"{total_periods}次"],
                ["平均月收益率", f"{total_return/total_periods:.2f}%" if total_periods > 0 else "N/A"]
            ]
            
            for key, value in performance:
                ws_overview[f'A{current_row}'] = key
                ws_overview[f'B{current_row}'] = value
                current_row += 1
        
        # 选股详情工作表
        ws_selection = wb.create_sheet("Stock_Selection")
        headers = ["期数", "选股日期", "调仓日期", "股票代码", "股票名称", "市值(亿)", "股价", "TTM_EPS", "TTM_PE", "权重"]
        
        for i, header in enumerate(headers, 1):
            ws_selection.cell(1, i, header)
        
        row = 2
        for pos in self.positions:
            for _, stock in pos['stocks'].iterrows():
                ws_selection.cell(row, 1, pos['period'])
                ws_selection.cell(row, 2, pos['selection_date'])
                ws_selection.cell(row, 3, pos['rebalance_date'])
                ws_selection.cell(row, 4, stock['stock_code'])
                ws_selection.cell(row, 5, stock['stock_name'])
                ws_selection.cell(row, 6, stock['market_cap_yi'])
                ws_selection.cell(row, 7, stock['close_price'])
                ws_selection.cell(row, 8, stock['ttm_eps'])
                ws_selection.cell(row, 9, stock['ttm_pe'])
                ws_selection.cell(row, 10, f"{stock['weight']*100:.1f}%")
                row += 1
        
        wb.save(f"{result_dir}/backtest_results.xlsx")
    
    def create_html_template(self, nav_data, params):
        """创建HTML图表模板"""
        
        # 准备数据
        dates = [item['date'] for item in nav_data]
        values = [item['nav'] for item in nav_data]
        
        # 计算业绩指标
        final_nav = values[-1] if values else 1.0
        years = (pd.to_datetime(self.end_date) - pd.to_datetime(self.start_date)).days / 365.25
        annual_return = (final_nav ** (1/years) - 1) * 100 if years > 0 else 0
        
        # 计算纵轴范围
        min_val = min(values) if values else 1.0
        max_val = max(values) if values else 1.0
        
        # 确保Y轴从合理的基准开始，通常策略图从0.5开始显示比较合理
        # 对于净值策略，Y轴应该包含1.0这个基准点
        y_min = max(0.5, min_val - 0.2)  # 最低0.5，或者稍低于最小值
        y_max = max_val + 0.5
        
        # 格式化日期为JavaScript可用格式
        js_dates = json.dumps([date.replace('-', '/') for date in dates])
        js_values = json.dumps([float(val) if hasattr(val, 'item') else val for val in values])
        
        
        html_content = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>策略回测图表</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-adapter-date-fns@3.0.0/dist/chartjs-adapter-date-fns.bundle.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-annotation@3.0.1/dist/chartjs-plugin-annotation.min.js"></script>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Microsoft YaHei', 'PingFang SC', 'SimHei', sans-serif;
            background: linear-gradient(135deg, #faf8f3 0%, #f5f0e8 100%);
            padding: 40px;
            min-height: 100vh;
        }}
        
        .container {{
            max-width: 1600px;
            margin: 0 auto;
            background: rgba(255,255,255,0.95);
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        
        .header {{
            background: linear-gradient(135deg, #ff7f50 0%, #ff6347 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        
        .header h1 {{
            font-size: 32px;
            font-weight: bold;
            margin-bottom: 8px;
        }}
        
        .header p {{
            font-size: 22px;
            opacity: 0.9;
            font-weight: 500;
        }}
        
        .dashboard {{
            display: grid;
            grid-template-columns: 1fr;
            gap: 30px;
            padding: 40px;
        }}
        
        .chart-section {{
            position: relative;
        }}
        
        .chart-container {{
            position: relative;
            height: 500px;
            background: white;
            border-radius: 15px;
            padding: 20px;
            box-shadow: 0 8px 25px rgba(0,0,0,0.08);
        }}
        
        
        .params-section {{
            background: linear-gradient(135deg, #f5f0e8 0%, #ede3d3 100%);
            border-radius: 15px;
            padding: 30px;
            border-left: 5px solid #ff7f50;
            box-shadow: 0 8px 25px rgba(0,0,0,0.08);
            display: grid;
            grid-template-columns: repeat(5, 1fr);
            gap: 20px;
        }}
        
        .params-section h3 {{
            color: #8b4513;
            font-size: 24px;
            margin-bottom: 25px;
            text-align: center;
            padding-bottom: 10px;
            border-bottom: 2px solid #deb887;
            grid-column: 1 / -1;
        }}
        
        .param-group {{
            margin-bottom: 25px;
        }}
        
        .param-group h4 {{
            color: #8b4513;
            font-size: 18px;
            margin-bottom: 12px;
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        
        .param-item {{
            color: #8b4513;
            font-size: 16px;
            line-height: 1.8;
            margin-left: 25px;
            position: relative;
        }}
        
        .param-item:before {{
            content: "•";
            color: #ff7f50;
            font-weight: bold;
            position: absolute;
            left: -15px;
        }}
        
        @media (max-width: 1200px) {{
            .params-section {{
                grid-template-columns: repeat(3, 1fr);
            }}
        }}
        
        @media (max-width: 800px) {{
            .params-section {{
                grid-template-columns: repeat(2, 1fr);
            }}
        }}
        
        @media (max-width: 600px) {{
            .params-section {{
                grid-template-columns: 1fr;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{self.strategy_name}净值走势图</h1>
            <p>最终净值: {final_nav:.2f} | 年化收益率: {annual_return:.2f}%</p>
        </div>
        
        <div class="dashboard">
            <div class="chart-section">
                <div class="chart-container">
                    <canvas id="navChart"></canvas>
                </div>
            </div>
            
            <div class="params-section">
                <h3>📊 策略参数配置</h3>
                
                <div class="param-group">
                    <h4>📊 基本参数</h4>
                    <div class="param-item">最低市值: {params['min_market_cap']}亿元</div>
                    <div class="param-item">选股数量: {params['stock_count']}只</div>
                    <div class="param-item">手续费率: {params['transaction_cost']*10000:.1f}‱</div>
                </div>
                
                <div class="param-group">
                    <h4>📅 回测时间</h4>
                    <div class="param-item">开始: {params['start_date']}</div>
                    <div class="param-item">结束: {params['end_date']}</div>
                </div>
                
                <div class="param-group">
                    <h4>🎯 选股逻辑</h4>
                    <div class="param-item">沪深主板股票</div>
                    <div class="param-item">市值≥{params['min_market_cap']}亿元</div>
                    <div class="param-item">按TTM PE升序选股</div>
                    <div class="param-item">等权重配置</div>
                </div>
                
                <div class="param-group">
                    <h4>⏰ 调仓规则</h4>
                    <div class="param-item">月度调仓</div>
                    <div class="param-item">月末选股</div>
                    <div class="param-item">次月首个交易日执行</div>
                </div>
                
                <div class="param-group">
                    <h4>💰 资金建议</h4>
                    <div class="param-item">建议资金: {params['stock_count']*2}万元以上</div>
                </div>
            </div>
        </div>
    </div>

    <script>
        // 等待DOM加载完成
        document.addEventListener('DOMContentLoaded', function() {{
            const ctx = document.getElementById('navChart').getContext('2d');
            
            const chart = new Chart(ctx, {{
                type: 'line',
                data: {{
                    labels: {js_dates},
                    datasets: [{{
                        label: '净值',
                        data: {js_values},
                        borderColor: '#ff7f50',
                        backgroundColor: 'rgba(255,127,80,0.1)',
                        borderWidth: 4,
                        pointBackgroundColor: '#ff6347',
                        pointBorderColor: 'white',
                        pointBorderWidth: 3,
                        pointRadius: 5,
                        pointHoverRadius: 7,
                        fill: true,
                        tension: 0.2
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{
                            display: false
                        }},
                        // 添加1.0基准线
                        annotation: {{
                            annotations: {{
                                baseLine: {{
                                    type: 'line',
                                    yMin: 1.0,
                                    yMax: 1.0,
                                    borderColor: '#ff6347',
                                    borderWidth: 2,
                                    borderDash: [5, 5],
                                    label: {{
                                        content: '基准线 1.0',
                                        enabled: true,
                                        position: 'end'
                                    }}
                                }}
                            }}
                        }},
                        tooltip: {{
                            backgroundColor: 'rgba(139,69,19,0.9)',
                            titleColor: 'white',
                            bodyColor: 'white',
                            borderColor: '#deb887',
                            borderWidth: 2,
                            cornerRadius: 8,
                            displayColors: false,
                            callbacks: {{
                                title: function(context) {{
                                    return '日期: ' + context[0].label;
                                }},
                                label: function(context) {{
                                    return '净值: ' + context.parsed.y.toFixed(4);
                                }}
                            }}
                        }}
                    }},
                    scales: {{
                        x: {{
                            grid: {{
                                color: 'rgba(210,180,140,0.3)',
                                lineWidth: 1
                            }},
                            ticks: {{
                                color: '#8b4513',
                                font: {{
                                    size: 12,
                                    weight: 'bold'
                                }},
                                maxRotation: 0,
                                stepSize: 1,
                                autoSkip: false,
                                callback: function(value, index, values) {{
                                    const totalLabels = this.chart.data.labels.length;
                                    
                                    // 计算要显示的索引：开始、结束和中间5个点（总共7个点）
                                    const step = Math.floor(totalLabels / 6);
                                    const showIndices = [0]; // 开始点
                                    
                                    for (let i = 1; i <= 5; i++) {{
                                        showIndices.push(Math.min(step * i, totalLabels - 1));
                                    }}
                                    showIndices.push(totalLabels - 1); // 结束点
                                    
                                    // 去重并排序
                                    const uniqueIndices = [...new Set(showIndices)].sort((a, b) => a - b);
                                    
                                    if (uniqueIndices.includes(index)) {{
                                        const labelValue = this.chart.data.labels[index];
                                        if (labelValue) {{
                                            const date = new Date(labelValue);
                                            const year = date.getFullYear();
                                            const month = String(date.getMonth() + 1).padStart(2, '0');
                                            return year + '-' + month;
                                        }}
                                    }}
                                    return '';
                                }}
                            }},
                            title: {{
                                display: true,
                                text: '日期',
                                color: '#8b4513',
                                font: {{
                                    size: 14,
                                    weight: 'bold'
                                }}
                            }}
                        }},
                        y: {{
                            min: {y_min:.1f},
                            max: {y_max:.1f},
                            grid: {{
                                color: 'rgba(210,180,140,0.3)',
                                lineWidth: 1
                            }},
                            ticks: {{
                                color: '#8b4513',
                                font: {{
                                    size: 12,
                                    weight: 'bold'
                                }},
                                callback: function(value) {{
                                    return value.toFixed(1);
                                }}
                            }},
                            title: {{
                                display: true,
                                text: '净值',
                                color: '#8b4513',
                                font: {{
                                    size: 14,
                                    weight: 'bold'
                                }}
                            }}
                        }}
                    }},
                    interaction: {{
                        intersect: false,
                        mode: 'index'
                    }},
                    // 禁用导出时动画，避免截图发生在动画过程中的中间帧
                    animation: false
                }}
            }});
            
            // 标记渲染完成
            window.chartReady = true;
        }});
    </script>
</body>
</html>
        """
        
        return html_content
    
    def generate_chart(self, result_dir):
        """生成策略净值图表"""
        if not self.nav_history:
            print("⚠️  无净值数据，跳过图表生成")
            return
        
        try:
            self._generate_html_chart(result_dir)
            print("📊 图表生成成功")
        except Exception as e:
            print(f"❌ 图表生成失败: {e}")
            raise
    
    def _generate_html_chart(self, result_dir):
        """HTML + Chart.js图表生成"""
        # 准备参数数据
        params = {
            'min_market_cap': self.min_market_cap,
            'stock_count': self.stock_count,
            'transaction_cost': self.transaction_cost,
            'start_date': self.start_date,
            'end_date': self.end_date
        }
        
        # 生成HTML内容
        html_content = self.create_html_template(self.nav_history, params)
        
        # 创建临时HTML文件 - 使用固定路径避免权限问题
        temp_html_path = f"{result_dir}/temp_chart.html"
        
        try:
            with open(temp_html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            # 使用Playwright生成截图
            self._screenshot_with_playwright(temp_html_path, f"{result_dir}/net_value_chart.png")
        finally:
            # 确保清理临时文件
            if os.path.exists(temp_html_path):
                os.unlink(temp_html_path)
    
    def _screenshot_with_playwright(self, html_path, output_path):
        """使用Playwright将HTML转换为图片"""
        with sync_playwright() as p:
            browser = None
            try:
                browser = p.chromium.launch(
                    headless=True,
                    args=['--no-sandbox', '--disable-dev-shm-usage']
                )
                page = browser.new_page()
                
                # 设置更长的超时时间
                page.set_default_navigation_timeout(60000)  # 60秒
                page.set_default_timeout(60000)
                
                # 设置页面大小
                page.set_viewport_size({"width": 1600, "height": 900})
                
                # 加载HTML文件，使用基本的domcontentloaded
                page.goto(f"file://{html_path}", wait_until="domcontentloaded")
                
                # 等待更长时间确保CDN资源加载
                page.wait_for_timeout(5000)
                
                # 等待图表渲染完成，增加超时时间
                try:
                    page.wait_for_function("window.chartReady === true", timeout=30000)
                except Exception:
                    print("⚠️  图表渲染检测超时，但继续尝试截图...")
                    # 再等待一下，可能图表仍在渲染
                    page.wait_for_timeout(3000)
                
                # 额外等待确保完全渲染
                page.wait_for_timeout(1000)
                
                # 截图
                page.screenshot(
                    path=output_path,
                    full_page=True
                )
                
                browser.close()
                
            except Exception as e:
                if browser is not None:
                    browser.close()
                raise e
    
    
    def generate_readme(self, result_dir):
        """生成README文件"""
        readme_content = f"""# {self.strategy_name} 回测报告

## 策略概述
- **策略名称**: {self.strategy_name}
- **回测期间**: {self.start_date} 至 {self.end_date}
- **选股范围**: 沪深主板市值≥{self.min_market_cap}亿股票
- **选股数量**: {self.stock_count}只
- **调仓频率**: 月度
- **手续费率**: {self.transaction_cost*10000:.1f}‱

## 核心业绩指标
"""
        
        if self.nav_history:
            final_nav = self.nav_history[-1]['nav']
            total_return = (final_nav - 1) * 100
            readme_content += f"""- **总收益率**: {total_return:.2f}%
- **最终净值**: {final_nav:.4f}
- **调仓次数**: {len(self.positions)}次

"""
        
        readme_content += """## 策略逻辑
1. 每月最后一个交易日，从沪深主板选择市值≥100亿的股票作为候选池
2. 计算所有候选股票的TTM PE（滚动12个月市盈率）
3. 按TTM PE升序排列，选择前10只股票
4. 等权重配置，每只股票权重10%
5. 下月第一个交易日执行调仓

## TTM PE计算方法
TTM EPS = 最近中报 + (上年年报 - 上年中报)
TTM PE = 当前股价 / TTM EPS

## 文件说明
- `backtest_results.xlsx`: 详细回测数据
- `net_value_chart.png`: 净值走势图
- `README.md`: 本报告文件
"""
        
        with open(f"{result_dir}/README.md", 'w', encoding='utf-8') as f:
            f.write(readme_content)
    
    def __del__(self):
        """清理资源"""
        if hasattr(self, 'conn'):
            self.conn.close()

def run_backtest(**kwargs):
    """策略入口函数"""
    strategy = LowTTMPEStrategy()
    strategy.run_backtest(**kwargs)

if __name__ == "__main__":
    # ==================== 策略参数设置 ====================
    # 以下参数可根据实际资金规模和需求调整
    
    # 基本策略参数
    MIN_MARKET_CAP = 100.0      # 最低市值要求(亿元) - 范围：50-500，建议100-200
    STOCK_COUNT = 10            # 选股数量 - 范围：5-30，建议10-20
    TRANSACTION_COST = 0.0001   # 交易手续费率 - 万分之一，可根据券商调整
    
    # 回测时间范围 (精确到月)
    START_DATE = "2020-01-01"   # 回测开始日期 - 格式：YYYY-MM-DD
    END_DATE = "2025-06-30"     # 回测结束日期 - 格式：YYYY-MM-DD
    
    # ==================== 不同资金规模建议 ====================
    # 2-5万元：STOCK_COUNT = 3-5,  MIN_MARKET_CAP = 50
    # 5-10万元：STOCK_COUNT = 5-8,  MIN_MARKET_CAP = 100  
    # 10-20万元：STOCK_COUNT = 10,   MIN_MARKET_CAP = 100 (推荐配置)
    # 20万元+：STOCK_COUNT = 15-20, MIN_MARKET_CAP = 150
    
    print("=" * 60)
    print("📈 主板低TTM PE轮动策略")
    print("=" * 60)
    print(f"📊 当前参数设置:")
    print(f"   最低市值: {MIN_MARKET_CAP}亿元")  
    print(f"   选股数量: {STOCK_COUNT}只")
    print(f"   手续费率: {TRANSACTION_COST*10000:.1f}‱")
    print(f"   回测期间: {START_DATE} 至 {END_DATE}")
    print(f"\n💡 建议资金规模: {STOCK_COUNT*2}万元以上")
    print("=" * 60)
    
    # 运行回测
    run_backtest(
        min_market_cap=MIN_MARKET_CAP,
        stock_count=STOCK_COUNT, 
        transaction_cost=TRANSACTION_COST,
        start_date=START_DATE,
        end_date=END_DATE
    )